import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import sys
import calendar
from collections import defaultdict

# Databas   e connection parameters
DB_CONFIG = {
    'host': 'HOSTNAME',
    'database': 'DATABASE_NAME',
    'user': 'USERNAME',
    'password': 'PASSWORD'
}

# Month names in German
MONTH_NAMES = {
    1: 'Jänner',
    2: 'Februar',
    3: 'März',
    4: 'April',
    5: 'Mai',
    6: 'Juni',
    7: 'Juli',
    8: 'August',
    9: 'September',
    10: 'Oktober',
    11: 'November',
    12: 'Dezember'
}

def get_database_data(year=2025):
    """Fetch all income/expense data from Postgres database for the year"""
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()

    query = """
        SELECT id, orderdate, who, position, income, expense, location, comment
        FROM incomeexpense
        WHERE EXTRACT(YEAR FROM orderdate) = %s
        ORDER BY orderdate, id
    """

    cursor.execute(query, (year,))
    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    return rows

def create_header_section(ws):
    """Create the fixed header section (rows 1-11) for a sheet"""

    # Define styles
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    cyan_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    purple_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")

    bold_font = Font(bold=True)
    white_font = Font(color="FFFFFF")

    # ===== ROW 1: Section Headers =====
    ws['A1'] = 'Zusammenfassung'
    ws['A1'].fill = red_fill
    ws.merge_cells('A1:B1')

    ws['D1'] = 'Fixkosten'
    ws['D1'].fill = red_fill
    ws.merge_cells('D1:E1')

    ws['G1'] = 'Fixkosten'
    ws['G1'].fill = red_fill
    ws.merge_cells('G1:H1')

    ws['J1'] = 'Allgemeine Ausgaben'
    ws['J1'].fill = red_fill
    ws.merge_cells('J1:K1')

    ws['M1'] = 'Allgemeine Ausgaben'
    ws['M1'].fill = red_fill
    ws.merge_cells('M1:N1')

    ws['P1'] = 'Passive monatliche(s) Einkommen/Ausgaben'
    ws['P1'].fill = red_fill
    ws.merge_cells('P1:R1')

    ws['T1'] = 'Jährliches Einkommen/Ausgaben'
    ws['T1'].fill = blue_fill
    ws['T1'].font = white_font
    ws.merge_cells('T1:U1')

    # ===== ROWS 2-11: Summary and Fixed Costs =====
    # Column A+B: Summary
    ws['A2'] = 'Einkommen'
    ws['B2'] = '=SUM(E14:E999)'
    ws['A2'].fill = green_fill
    ws['B2'].number_format = '#,##0.00 [$€]'

    ws['A3'] = 'Ausgaben'
    ws['B3'] = '=SUM(F14:F999)'
    ws['A3'].fill = cyan_fill
    ws['B3'].number_format = '#,##0.00 [$€]'

    ws['A4'] = 'Erspartes'
    ws['B4'] = '=B2+B3'
    ws['A4'].fill = yellow_fill
    ws['B4'].number_format = '#,##0.00 [$€]'

    ws['A5'] = 'Basis für Essen/Monat'
    ws['B5'] = 600.00
    ws['A5'].fill = yellow_fill
    ws['B5'].number_format = '#,##0.00 [$€]'

    ws['A6'] = 'Verbleibendes Essensgeld/Monat'
    ws['B6'] = '=B5+K5'
    ws['A6'].fill = purple_fill
    ws['B6'].number_format = '#,##0.00 [$€]'

    ws['A7'] = 'Essensgeld Durchschnitt/Tag bis Monatsende'
    ws['B7'] = '=B6/(EOMONTH(TODAY(),0)-TODAY()+1)'
    ws['A7'].fill = purple_fill
    ws['B7'].number_format = '#,##0.00 [$€]'

    # Column D+E: Fixkosten Group 1
    fixkosten_1 = [
        ('Bank', 'D2'),
        ('Betriebsratsumlage', 'D3'),
        ('Betriebskosten Garagenplatz A1/12', 'D4'),
        ('Betriebskosten Garagenplatz A3/17', 'D5'),
        ('Internet', 'D6'),
        ('OBS-Haushaltsabgabe', 'D7')
    ]

    for label, cell in fixkosten_1:
        ws[cell] = label
        value_cell = cell.replace('D', 'E')
        # SUMIF formula: sum income + expenses where position matches
        ws[value_cell] = f'=SUMIF($D$14:$D$999,{cell},$E$14:$E$999)+SUMIF($D$14:$D$999,{cell},$F$14:$F$999)'
        ws[value_cell].number_format = '#,##0.00 [$€]'

    # Column G+H: Fixkosten Group 2
    fixkosten_2 = [
        ('Reparaturrücklage Garagenplatz A1/12', 'G2'),
        ('Reparaturrücklage Garagenplatz A3/17', 'G3'),
        ('Strom', 'G4'),
        ('Telefon', 'G5'),
        ('Versicherung', 'G6'),
        ('Wasser/Heizung', 'G7')
    ]

    for label, cell in fixkosten_2:
        ws[cell] = label
        value_cell = cell.replace('G', 'H')
        # SUMIF formula: sum income + expenses where position matches
        ws[value_cell] = f'=SUMIF($D$14:$D$999,{cell},$E$14:$E$999)+SUMIF($D$14:$D$999,{cell},$F$14:$F$999)'
        ws[value_cell].number_format = '#,##0.00 [$€]'

    # Column J+K: Allgemeine Ausgaben Group 1
    allgemeine_ausgaben_1 = [
        ('Arbeitssuche', 'J2'),
        ('Auto', 'J3'),
        ('Digitale Arbeitsmittel', 'J4'),
        ('Essen', 'J5'),
        ('Fachliteratur', 'J6'),
        ('Fortbildung', 'J7')
    ]

    for label, cell in allgemeine_ausgaben_1:
        ws[cell] = label
        value_cell = cell.replace('J', 'K')
        # SUMIF formula: sum income + expenses where position matches
        ws[value_cell] = f'=SUMIF($D$14:$D$999,{cell},$E$14:$E$999)+SUMIF($D$14:$D$999,{cell},$F$14:$F$999)'
        ws[value_cell].number_format = '#,##0.00 [$€]'

    # Column M+N: Allgemeine Ausgaben Group 2
    allgemeine_ausgaben_2 = [
        ('Fun', 'M2'),
        ('Kammer', 'M3'),
        ('Klimaanlage', 'M4'),
        ('Medizin', 'M5'),
        ('Shop', 'M6'),
        ('Steuerberater', 'M7')
    ]

    for label, cell in allgemeine_ausgaben_2:
        ws[cell] = label
        value_cell = cell.replace('M', 'N')
        # SUMIF formula: sum income + expenses where position matches
        ws[value_cell] = f'=SUMIF($D$14:$D$999,{cell},$E$14:$E$999)+SUMIF($D$14:$D$999,{cell},$F$14:$F$999)'
        ws[value_cell].number_format = '#,##0.00 [$€]'

    # Column P+Q+R: Passive monthly income/expenses
    # Headers
    ws['Q2'] = 'Einnahmen'
    ws['Q2'].font = bold_font
    ws['R2'] = 'Ausgaben'
    ws['R2'].font = bold_font

    # Location rows
    passive = [
        ('Hollgasse 1/1', 'P3'),
        ('Hollgasse 1/54', 'P4'),
        ('Stipcakgasse 8', 'P5')
    ]

    for label, cell in passive:
        ws[cell] = label
        value_cell_q = cell.replace('P', 'Q')
        value_cell_r = cell.replace('P', 'R')
        ws[value_cell_q] = 0
        ws[value_cell_r] = 0
        ws[value_cell_q].number_format = '0'
        ws[value_cell_r].number_format = '0'

    # Column T+U: Jährliches Einkommen/Ausgaben (Yearly income/expense)
    ws['T2'] = 'Einkommen'
    ws['U2'] = '=SUM(Jänner:Dezember!B2)'
    ws['T2'].fill = green_fill
    ws['U2'].number_format = '#,##0.00 [$€]'

    ws['T3'] = 'Ausgaben'
    ws['U3'] = '=SUM(Jänner:Dezember!B3)'
    ws['T3'].fill = cyan_fill
    ws['U3'].number_format = '#,##0.00 [$€]'

    ws['T4'] = 'Ersparnis'
    ws['U4'] = '=SUM(Jänner:Dezember!B4)'
    ws['T4'].fill = yellow_fill
    ws['U4'].number_format = '#,##0.00 [$€]'

    ws['T5'] = '% Erspartes'
    ws['U5'] = '=U4/U2'
    ws['T5'].fill = yellow_fill
    ws['U5'].number_format = '0.00%'

    ws['T6'] = '% Ausgaben'
    ws['U6'] = '=U3/U2'
    ws['T6'].fill = yellow_fill
    ws['U6'].number_format = '0.00%'

    # ===== ROW 12: Data Headers =====
    headers = [
        'wochentag',
        'datum',
        'person',
        'position',
        'einnahmen',
        'ausgaben',
        'location',
        'comment'
    ]

    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col_idx)
        cell.value = header
        cell.font = bold_font

    # Set column widths
    column_widths = {
        'A': 15,  # wochentag / Zusammenfassung
        'B': 15,  # datum / values
        'C': 10,  # person
        'D': 35,  # position / Fixkosten labels
        'E': 15,  # einnahmen / Fixkosten values
        'F': 15,  # ausgaben
        'G': 35,  # location / Fixkosten labels
        'H': 15,  # comment / Fixkosten values
        'I': 3,   # separator
        'J': 25,  # Allgemeine Ausgaben labels
        'K': 15,  # Allgemeine Ausgaben values
        'L': 3,   # separator
        'M': 20,  # Allgemeine Ausgaben labels
        'N': 15,  # Allgemeine Ausgaben values
        'O': 3,   # separator
        'P': 20,  # Passive income location labels
        'Q': 12,  # Passive Einnahmen values
        'R': 12,  # Passive Ausgaben values
        'S': 3,   # separator
        'T': 20,  # Yearly summary labels
        'U': 15,  # Yearly summary values
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze rows 1-12 (header section)
    ws.freeze_panes = 'A13'

    # Add borders around specific ranges
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    border_ranges = [
        ('A1', 'B7'),
        ('D1', 'E7'),
        ('G1', 'H7'),
        ('J1', 'K7'),
        ('M1', 'N7'),
        ('P1', 'R7'),
        ('T1', 'U7')
    ]

    for start_cell, end_cell in border_ranges:
        # Parse cell coordinates
        from openpyxl.utils import column_index_from_string, get_column_letter

        start_col = column_index_from_string(start_cell[0])
        start_row = int(start_cell[1])
        end_col = column_index_from_string(end_cell[0])
        end_row = int(end_cell[1:])

        # Apply borders to all cells in the range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)

                # Determine which borders this cell needs
                left = Side(style='thin') if col == start_col else None
                right = Side(style='thin') if col == end_col else None
                top = Side(style='thin') if row == start_row else None
                bottom = Side(style='thin') if row == end_row else None

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def add_month_data(ws, month_rows):
    """Add transaction data for a specific month to a sheet"""
    current_row = 13
    current_date = None

    for row in month_rows:
        db_id, orderdate, who, position, income, expense, location, comment = row

        # Check if we need a new date header
        if current_date != orderdate:
            current_date = orderdate
            day_name = calendar.day_name[orderdate.weekday()]

            # Add day of week
            ws.cell(row=current_row, column=1).value = day_name
            ws.cell(row=current_row, column=2).value = orderdate
            ws.cell(row=current_row, column=2).number_format = 'dd.mm.yyyy'
            current_row += 1

        # Add transaction row
        ws.cell(row=current_row, column=3).value = who
        ws.cell(row=current_row, column=4).value = position

        # Column E: einnahmen (income only)
        if income and float(income) > 0:
            ws.cell(row=current_row, column=5).value = float(income)
            ws.cell(row=current_row, column=5).number_format = '#,##0.00 [$€]'

        # Column F: ausgaben (expenses only - with minus sign)
        if expense and float(expense) > 0:
            ws.cell(row=current_row, column=6).value = -float(expense)
            ws.cell(row=current_row, column=6).number_format = '#,##0.00 [$€]'

        # Column G: location
        ws.cell(row=current_row, column=7).value = location

        # Column H: comment
        ws.cell(row=current_row, column=8).value = comment if comment else ''

        current_row += 1

def generate_excel(db_rows, year=2025, output_file=None):
    """Generate Excel file with 12 sheets (one per month)"""

    if output_file is None:
        output_file = f'einnahmen_ausgaben_{year}.xlsx'

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # Group data by month
    monthly_data = defaultdict(list)
    for row in db_rows:
        db_id, orderdate, who, position, income, expense, location, comment = row
        month = orderdate.month
        monthly_data[month].append(row)

    # Create a sheet for each month
    for month in range(1, 13):
        month_name = MONTH_NAMES[month]
        ws = wb.create_sheet(title=month_name)

        # Add header section
        create_header_section(ws)

        # Add data for this month (if any)
        if month in monthly_data:
            add_month_data(ws, monthly_data[month])

    # Save workbook
    wb.save(output_file)
    return output_file

def main():
    """Main function to export general income/expense report"""
    try:
        # Get year from command line or use current year
        if len(sys.argv) > 1:
            try:
                year = int(sys.argv[1])
            except ValueError:
                print(f"Invalid year: {sys.argv[1]}")
                print("Usage: python3 einnahmen_ausgaben.py [year]")
                return
        else:
            year = datetime.now().year

        print(f"Generating general income/expense report for year: {year}")
        print(f"Fetching data from database...")
        db_rows = get_database_data(year=year)

        if not db_rows:
            print(f"No data found for {year}")
            return

        print(f"Found {len(db_rows)} total records")

        output_file = f'einnahmen_ausgaben_{year}.xlsx'
        print(f"\nGenerating Excel file with 12 monthly sheets: {output_file}")
        generate_excel(db_rows, year=year, output_file=output_file)

        print(f"✓ Excel file created successfully: {output_file}")
        print(f"  ✓ Created 12 sheets (one for each month)")
        print(f"  ✓ Each sheet has fixed header section and monthly data")
        print(f"\nTo download from container:")
        print(f"docker cp $(docker-compose ps -q web):/app/{output_file} ./exports/{output_file}")

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()