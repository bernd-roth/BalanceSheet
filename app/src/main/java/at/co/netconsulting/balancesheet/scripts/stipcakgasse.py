import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from collections import defaultdict
import sys
import os

# Database connection parameters
DB_CONFIG = {
    'host': 'HOSTNAME',
    'database': 'DATABASE_NAME',
    'user': 'USERNAME',
    'password': 'PASSWORD'
}

# Month names in German
MONTH_NAMES = {
    1: 'Jänner',
    2: 'Februar',
    3: 'März',
    4: 'April',
    5: 'Mai',
    6: 'Juni',
    7: 'Juli',
    8: 'August',
    9: 'September',
    10: 'Oktober',
    11: 'November',
    12: 'Dezember'
}

# Template categories for parking lots (ordered by column appearance)
TEMPLATE_CATEGORIES = [
    'mieteinkommen',
    'betriebskosten_a3_17',
    'reparatur_a3_17',
    'betriebskosten_a1_12',
    'reparatur_a1_12'
]

def get_category_column(position, comment):
    """
    Determine which column based on position field for parking lots
    Returns the column name or None if not parking-related
    """
    position_lower = position.lower().strip() if position else ''

    # Direct mapping for parking lot positions matching Position enum
    # Database stores displayName values, not enum names
    parking_positions = {
        'vermietung garage': 'mieteinkommen',
        'betriebskosten garagenplatz a1/12': 'betriebskosten_a1_12',
        'betriebskosten garagenplatz a3/17': 'betriebskosten_a3_17',
        'reparaturrücklage garagenplatz a1/12': 'reparatur_a1_12',
        'reparaturrücklage garagenplatz a3/17': 'reparatur_a3_17',
    }

    # Check position match only
    return parking_positions.get(position_lower, None)

def get_database_data(location='Stipcakgasse 8', year=2025, taxable=None):
    """Fetch parking lot data from Postgres database

    Args:
        location: The location name (e.g., 'Stipcakgasse 8')
        year: The year to filter by
        taxable: If True, filter for taxable entries. If False, filter for non-taxable entries.
                 If None, return all entries for the location.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()

    if taxable is not None:
        query = """
            SELECT id, orderdate, position, income, expense, comment, taxable
            FROM incomeexpense
            WHERE location LIKE %s
            AND EXTRACT(YEAR FROM orderdate) = %s
            AND taxable = %s
            ORDER BY orderdate, id
        """
        cursor.execute(query, (location + '%', year, taxable))
    else:
        query = """
            SELECT id, orderdate, position, income, expense, comment, taxable
            FROM incomeexpense
            WHERE location LIKE %s
            AND EXTRACT(YEAR FROM orderdate) = %s
            ORDER BY orderdate, id
        """
        cursor.execute(query, (location + '%', year))

    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    return rows


def aggregate_data_by_month(db_rows):
    """Aggregate database entries by month and category"""
    monthly_data = defaultdict(lambda: defaultdict(lambda: {'amount': 0.0, 'entries': []}))

    for row in db_rows:
        # Handle both old (6 fields) and new (7 fields with taxable) row formats
        if len(row) == 7:
            db_id, orderdate, position, income, expense, comment, taxable = row
        else:
            db_id, orderdate, position, income, expense, comment = row
            taxable = False  # Default for old data

        category = get_category_column(position, comment)
        if category is None:
            continue

        month = orderdate.month
        amount = float(income) if float(income) > 0 else -float(expense)

        monthly_data[month][category]['amount'] += amount
        monthly_data[month][category]['entries'].append({
            'date': orderdate,
            'description': position,
            'amount': amount,
            'comment': comment if comment else ''
        })

    return monthly_data


def generate_excel(monthly_data, location='Stipcakgasse 8/1', year=2025, output_file=None):
    """Generate Excel file for parking lots"""
    
    if output_file is None:
        output_file = f'stipcakgasse_{year}.xlsx'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "parking_stipcakgasse"
    
    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Row 1: Title with yellow background
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f"Einnahmen und \nAusgaben\nfür das Jahr {year}\nParkplätze Stipcakgasse 8/1"
    title_cell.fill = yellow_fill
    title_cell.alignment = centered_alignment
    title_cell.font = Font(bold=True, size=12)
    ws.row_dimensions[1].height = 75

    # Row 2: Column headers
    headers = [
        'rechnungsnummer',
        'datum',
        'artikelbeschreibung',
        'ein / aus',
        'mieteinkommen',
        'betriebs-\nkosten\na3/17',
        'reparatur-\nrücklage\na3/17',
        'betriebs-\nkosten\na1/12',
        'reparatur-\nrücklage\na1/12',
        'comment'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.alignment = header_alignment
        cell.font = Font(bold=True)
        cell.border = thin_border
    
    ws.row_dimensions[2].height = 45
    
    # Freeze first 2 rows
    ws.freeze_panes = 'A3'
    
    # Set column widths
    column_widths = {
        'A': 18,  # rechnungsnummer
        'B': 12,  # datum
        'C': 25,  # artikelbeschreibung
        'D': 12,  # ein/aus
        'E': 13,  # mieteinkommen
        'F': 12,  # betriebskosten a3/17
        'G': 12,  # reparaturrücklage a3/17
        'H': 12,  # betriebskosten a1/12
        'I': 12,  # reparaturrücklage a1/12
        'J': 30,  # comment
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Initialize totals
    totals = {
        'ein_aus': 0.0,
        'mieteinkommen': 0.0,
        'betriebskosten_a3_17': 0.0,
        'reparatur_a3_17': 0.0,
        'betriebskosten_a1_12': 0.0,
        'reparatur_a1_12': 0.0,
    }

    invoice_number = 1
    current_row = 3

    # Column mapping
    category_columns = {
        'mieteinkommen': 5,          # E
        'betriebskosten_a3_17': 6,   # F
        'reparatur_a3_17': 7,        # G
        'betriebskosten_a1_12': 8,   # H
        'reparatur_a1_12': 9,        # I
    }
    
    # Write data for each month
    for month in range(1, 13):
        month_name = MONTH_NAMES[month]
        
        # Month header row
        ws.cell(row=current_row, column=2).value = month_name
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        current_row += 1
        
        categories_in_month = monthly_data.get(month, {})

        # Write entries for each category
        for category in TEMPLATE_CATEGORIES:
            if category in categories_in_month:
                data = categories_in_month[category]

                for entry in data['entries']:
                    entry_amount = entry['amount']

                    # Invoice number
                    invoice_cell = ws.cell(row=current_row, column=1)
                    invoice_cell.value = invoice_number
                    invoice_cell.number_format = '000'
                    invoice_number += 1

                    # Date
                    date_cell = ws.cell(row=current_row, column=2)
                    date_cell.value = entry['date']
                    date_cell.number_format = 'dd.mm.yyyy'

                    # Description (Column C)
                    ws.cell(row=current_row, column=3).value = entry['description']

                    # ein/aus (Column D)
                    ein_aus_cell = ws.cell(row=current_row, column=4)
                    ein_aus_cell.value = entry_amount
                    ein_aus_cell.number_format = '#,##0.00 [$€-1]'
                    totals['ein_aus'] += entry_amount

                    # Category column (E-I)
                    if category in category_columns:
                        col_idx = category_columns[category]
                        amount_cell = ws.cell(row=current_row, column=col_idx)
                        amount_cell.value = entry_amount
                        amount_cell.number_format = '#,##0.00 [$€-1]'
                        totals[category] += entry_amount

                    # Column J: Comment
                    ws.cell(row=current_row, column=10).value = entry.get('comment', '')

                    current_row += 1
            # Skip empty rows - no else clause
    
    # Add totals row
    ws.cell(row=current_row, column=1).value = "summe"
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    
    # Total ein/aus
    total_cell = ws.cell(row=current_row, column=4)
    total_cell.value = totals['ein_aus']
    total_cell.number_format = '#,##0.00 [$€-1]'
    total_cell.font = Font(bold=True)
    
    # Category totals
    for category, col_idx in category_columns.items():
        if totals[category] != 0:
            total_cat_cell = ws.cell(row=current_row, column=col_idx)
            total_cat_cell.value = totals[category]
            total_cat_cell.number_format = '#,##0.00 [$€-1]'
            total_cat_cell.font = Font(bold=True)
    
    wb.save(output_file)

    # Fix file permissions (make it readable/writable for user and group)
    try:
        os.chmod(output_file, 0o664)
    except Exception as e:
        print(f"Warning: Could not set file permissions: {e}")

    return output_file

def main():
    """Main function to export parking lot data"""
    try:
        location = 'Stipcakgasse 8'

        # Get year from command line or use current year
        if len(sys.argv) > 1:
            try:
                year = int(sys.argv[1])
            except ValueError:
                print(f"Invalid year: {sys.argv[1]}")
                print("Usage: python3 stipcakgasse.py [year] [taxable]")
                return
        else:
            year = datetime.now().year

        # Get taxable filter from command line argument (optional)
        # Options: 'true', 'false', or omit for all entries
        taxable_filter = None
        if len(sys.argv) > 2:
            taxable_arg = sys.argv[2].lower()
            if taxable_arg == 'true':
                taxable_filter = True
            elif taxable_arg == 'false':
                taxable_filter = False
            # else: None (all entries)

        print(f"Generating report for: {location}, year: {year}, taxable: {taxable_filter}")
        print(f"Fetching data from database...")
        db_rows = get_database_data(location=location, year=year, taxable=taxable_filter)
        
        if not db_rows:
            print(f"No data found for {location} in {year}")
            return
        
        print(f"Found {len(db_rows)} total records")

        # Debug: Show taxable status of fetched entries
        print("\nDEBUG - Fetched entries with taxable status:")
        for row in db_rows[:10]:  # Show first 10 entries
            if len(row) == 7:
                db_id, orderdate, position, income, expense, comment, taxable = row
                print(f"  ID {db_id}: {position} | {expense}€ | taxable={taxable}")
            else:
                db_id, orderdate, position, income, expense, comment = row
                print(f"  ID {db_id}: {position} | {expense}€ | taxable=MISSING")

        print("Aggregating parking lot data by month and category...")
        monthly_data = aggregate_data_by_month(db_rows)
        
        parking_entries = sum(len(cat['entries']) for month_data in monthly_data.values() 
                            for cat in month_data.values())
        print(f"Found {parking_entries} parking-related entries")
        
        print("\nMonthly breakdown:")
        for month in sorted(monthly_data.keys()):
            month_name = MONTH_NAMES.get(month, f"Month {month}")
            categories = list(monthly_data.keys())
            total_month = sum(cat['amount'] for cat in monthly_data[month].values())
            print(f"  {month_name}: {len(categories)} categories, Total: {total_month:.2f} €")
        
        output_file = f'stipcakgasse_{year}.xlsx'
        print(f"\nGenerating Excel file: {output_file}")
        generate_excel(monthly_data, location=location, year=year, output_file=output_file)
        
        print(f"✓ Excel file created successfully: {output_file}")
        print(f"\nTo download from container:")
        print(f"docker cp $(docker-compose ps -q web):/app/{output_file} ./exports/{output_file}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()