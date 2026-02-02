import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import sys

# Database connection parameters
DB_CONFIG = {
    'host': 'HOSTNAME',
    'database': 'DATABASE_NAME',
    'user': 'USERNAME',
    'password': 'PASSWORD'
}

# Month names in German
MONTH_NAMES = {
    1: 'Jänner',
    2: 'Februar',
    3: 'März',
    4: 'April',
    5: 'Mai',
    6: 'Juni',
    7: 'Juli',
    8: 'August',
    9: 'September',
    10: 'Oktober',
    11: 'November',
    12: 'Dezember'
}

def get_tax_category_column(position, comment, export_to='auto'):
    """
    Map database position to Austrian tax category column
    Returns column name or None if not applicable

    Args:
        position: The position/category name
        comment: Additional comment (currently unused)
        export_to: Export routing control ('auto', 'hollgasse', 'arbeitnehmerveranlagung', 'both', 'none')
    """
    # Check export_to field first
    if export_to == 'none':
        return None
    if export_to == 'hollgasse':
        return None  # This entry should only go to hollgasse, not arbeitnehmerveranlagung
    # If export_to is 'auto', 'arbeitnehmerveranlagung', or 'both', continue with position-based mapping

    position_lower = position.lower().strip() if position else ''
    comment_lower = comment.lower().strip() if comment else ''

    # Map positions to tax columns - exact match mapping
    # The key is what's in the database, value is the column category
    position_mapping = {
        'fortbildung': 'fortbildung',
        'fachliteratur': 'fachliteratur',
        'kammer': 'kammer',
        'medizin': 'medizin',
        'arbeitssuche': 'arbeitssuche',
        'strom': 'strom',
        'betriebsratsumlage': 'betriebsratsumlage',
        'homeoffice-pauschale': 'homeoffice_pauschale',
        'homeoffice_pauschale': 'homeoffice_pauschale',
        'steuerberater': 'steuerberater',
        'digitale arbeitsmittel': 'digitale_arbeitsmittel',
        'digitale_arbeitsmittel': 'digitale_arbeitsmittel',
        'versicherung': 'versicherung',
        'rechts -und beratungskosten': 'rechts_und_beratungskosten',
        'rechts_und_beratungskosten': 'rechts_und_beratungskosten',
        'internet': 'digitale_arbeitsmittel'  # Internet can be mapped to digitale_arbeitsmittel for personal tax
    }

    # For export_to='arbeitnehmerveranlagung' or 'both', force inclusion
    if export_to in ['arbeitnehmerveranlagung', 'both']:
        # Try to map the position, or use a default category if it doesn't match
        return position_mapping.get(position_lower, 'steuerberater')  # Default to 'steuerberater' for forced entries

    # For export_to='auto', use position-based mapping
    # Exclude rental-related positions when in auto mode
    rental_positions = [
        'mieteinkommen', 'garage a3/17', 'garage a1/12',
        'reparaturrücklage garage a3/17', 'reparaturrücklage garage a1/12',
        'betriebskosten garage a3/17', 'betriebskosten garage a1/12',
        'reparaturrücklage garagenplatz a3/17', 'reparaturrücklage garagenplatz a1/12',
        'betriebskosten garagenplatz a3/17', 'betriebskosten garagenplatz a1/12',
        'wasser/heizung', 'haushaltsversicherung', 'hausverwaltung',
        'internet', 'klimaanlage', 'obs haushaltsabgabe',
        'rechtsschutzversicherung', 'vermietung garage', 'bank',
        'auto', 'essen', 'einkommen', 'shop', 'telefon'
    ]

    if position_lower in rental_positions:
        return None

    # Return the mapped category if position matches
    return position_mapping.get(position_lower, None)


def get_database_data(person='Bernd', year=2025):
    """Fetch personal tax data from Postgres database"""
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()

    query = """
        SELECT id, orderdate, position, income, expense, comment, export_to
        FROM incomeexpense
        WHERE who = %s
        AND EXTRACT(YEAR FROM orderdate) = %s
        ORDER BY orderdate, id
    """

    cursor.execute(query, (person, year))
    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    return rows


def aggregate_data_by_month(db_rows):
    """Aggregate entries by month and tax category"""
    monthly_data = defaultdict(lambda: defaultdict(lambda: {'amount': 0.0, 'entries': []}))

    for row in db_rows:
        # Handle different row formats (old and new with export_to field)
        if len(row) == 7:
            db_id, orderdate, position, income, expense, comment, export_to = row
        else:
            db_id, orderdate, position, income, expense, comment = row
            export_to = 'auto'  # Default for data without export_to field

        category = get_tax_category_column(position, comment, export_to)
        if category is None:
            continue

        month = orderdate.month
        income_val = float(income) if income else 0.0
        expense_val = float(expense) if expense else 0.0

        # Determine the amount for ja/nein% and ansetzbar
        # Expenses get a negative sign
        if income_val != 0:
            amount = income_val
        elif expense_val != 0:
            amount = -expense_val  # Negative sign for expenses
        else:
            amount = 0.0

        monthly_data[month][category]['amount'] += amount
        monthly_data[month][category]['entries'].append({
            'date': orderdate,
            'description': position,
            'comment': comment if comment else '',
            'income': income_val,
            'expense': expense_val,
            'amount': amount
        })

    return monthly_data

def generate_excel(monthly_data, person='Bernd', year=2025, output_file=None):
    """Generate Excel file for Arbeitnehmerveranlagung"""

    if output_file is None:
        output_file = f'{person.lower()}_arbeitnehmerveranlagung_{year}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = f"{person}_ANV_{year}"

    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_alignment = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Title with yellow background
    ws.merge_cells('A1:S1')
    title_cell = ws['A1']
    title_cell.value = f"Arbeitnehmerveranlagung {person}\nfür das Jahr {year}"
    title_cell.fill = yellow_fill
    title_cell.alignment = centered_alignment
    title_cell.font = Font(bold=True, size=12)
    ws.row_dimensions[1].height = 45

    # Row 2: Column headers - added 'prozent' and 'ansetzbar' after 'betrag'
    headers = [
        'rechnungsnummer',
        'datum',
        'artikelbeschreibung',
        'betrag',
        'prozent',
        'ansetzbar',
        'fortbildung',
        'fachliteratur',
        'kammer',
        'medizin',
        'arbeitssuche',
        'strom',
        'betriebsrats-\numlage',
        'homeoffice-\npauschale',
        'steuerberater',
        'digitale\narbeitsmittel',
        'versicherung',
        'rechts -und\nberatungskosten',
        'comment'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.alignment = header_alignment
        cell.font = Font(bold=True, size=9)
        cell.border = thin_border

    ws.row_dimensions[2].height = 45

    # Freeze first 2 rows
    ws.freeze_panes = 'A3'

    # Set column widths
    column_widths = {
        'A': 18,  # rechnungsnummer
        'B': 12,  # datum
        'C': 25,  # artikelbeschreibung
        'D': 12,  # betrag
        'E': 10,  # prozent
        'F': 12,  # ansetzbar
        'G': 10,  # fortbildung
        'H': 12,  # fachliteratur
        'I': 10,  # kammer
        'J': 10,  # medizin
        'K': 12,  # arbeitssuche
        'L': 10,  # strom
        'M': 12,  # betriebsratsumlage
        'N': 12,  # homeoffice-pauschale
        'O': 12,  # steuerberater
        'P': 12,  # digitale arbeitsmittel
        'Q': 12,  # versicherung
        'R': 14,  # rechts -und beratungskosten
        'S': 30,  # comment
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    current_row = 3
    first_data_row = 3  # Track the first row with data for SUM formulas

    # Column mapping - maps category to Excel column number
    # Shifted by 2 columns due to new 'prozent' and 'ansetzbar' columns
    position_to_column = {
        'fortbildung': 7,                  # G
        'fachliteratur': 8,                # H
        'kammer': 9,                       # I
        'medizin': 10,                     # J
        'arbeitssuche': 11,                # K
        'strom': 12,                       # L
        'betriebsratsumlage': 13,          # M
        'homeoffice_pauschale': 14,        # N
        'steuerberater': 15,               # O
        'digitale_arbeitsmittel': 16,      # P
        'versicherung': 17,                # Q
        'rechts_und_beratungskosten': 18,  # R
    }

    invoice_number = 1

    # Write data for each month
    for month in range(1, 13):
        month_name = MONTH_NAMES[month]

        categories_in_month = monthly_data.get(month, {})

        if not categories_in_month:
            continue

        # Month header row
        ws.cell(row=current_row, column=2).value = month_name
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        current_row += 1

        # Write entries
        for category, data in categories_in_month.items():
            for entry in data['entries']:
                # Update first_data_row if this is the first entry
                if invoice_number == 1:
                    first_data_row = current_row
                # Rechnungsnummer
                invoice_cell = ws.cell(row=current_row, column=1)
                invoice_cell.value = invoice_number
                invoice_cell.number_format = '000'
                invoice_number += 1

                # Date
                date_cell = ws.cell(row=current_row, column=2)
                date_cell.value = entry['date']
                date_cell.number_format = 'dd.mm.yyyy'

                # Artikelbeschreibung (position from DB)
                ws.cell(row=current_row, column=3).value = entry['description']

                # Betrag (column D = 4) - the amount (negative for expenses)
                betrag_cell = ws.cell(row=current_row, column=4)
                betrag_cell.value = entry['amount']
                betrag_cell.number_format = '#,##0.00 [$€-1]'

                # Prozent (column E = 5) - default to 100%
                prozent_cell = ws.cell(row=current_row, column=5)
                prozent_cell.value = 100
                prozent_cell.number_format = '0'

                # Ansetzbar (column F = 6) - formula: =(D{row}/100)*E{row}
                ansetzbar_cell = ws.cell(row=current_row, column=6)
                ansetzbar_cell.value = f'=(D{current_row}/100)*E{current_row}'
                ansetzbar_cell.number_format = '#,##0.00 [$€-1]'

                # Map to specific column based on category
                # Use formula reference to ansetzbar column
                if category in position_to_column:
                    col_idx = position_to_column[category]
                    amount_cell = ws.cell(row=current_row, column=col_idx)
                    amount_cell.value = f'=F{current_row}'
                    amount_cell.number_format = '#,##0.00 [$€-1]'

                # Comment (column S = 19)
                ws.cell(row=current_row, column=19).value = entry.get('comment', '')

                current_row += 1

    # Add totals row
    ws.cell(row=current_row, column=2).value = "SUMME"
    ws.cell(row=current_row, column=2).font = Font(bold=True)

    # Calculate the last data row (current_row - 1)
    last_data_row = current_row - 1

    # Total Betrag (column D = 4) - use SUM formula
    total_cell = ws.cell(row=current_row, column=4)
    total_cell.value = f'=SUM(D{first_data_row}:D{last_data_row})'
    total_cell.number_format = '#,##0.00 [$€-1]'
    total_cell.font = Font(bold=True)

    # Total Ansetzbar (column F = 6) - use SUM formula
    total_ansetzbar_cell = ws.cell(row=current_row, column=6)
    total_ansetzbar_cell.value = f'=SUM(F{first_data_row}:F{last_data_row})'
    total_ansetzbar_cell.number_format = '#,##0.00 [$€-1]'
    total_ansetzbar_cell.font = Font(bold=True)

    # Category totals - use SUM formulas
    for category, col_idx in position_to_column.items():
        col_letter = get_column_letter(col_idx)
        total_cat_cell = ws.cell(row=current_row, column=col_idx)
        total_cat_cell.value = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
        total_cat_cell.number_format = '#,##0.00 [$€-1]'
        total_cat_cell.font = Font(bold=True)

    wb.save(output_file)
    return output_file

def main():
    """Main function to export Arbeitnehmerveranlagung data"""
    try:
        # Get person from command line (Bernd or Julia)
        if len(sys.argv) < 2:
            print("Usage: python3 arbeitnehmerveranlagung.py <person> [year]")
            print("Example: python3 arbeitnehmerveranlagung.py Bernd")
            print("Example: python3 arbeitnehmerveranlagung.py Julia 2025")
            return

        person = sys.argv[1].capitalize()

        if person not in ['Bernd', 'Julia']:
            print(f"Invalid person: {person}. Must be 'Bernd' or 'Julia'")
            return

        # Get year from command line or use current year
        if len(sys.argv) > 2:
            try:
                year = int(sys.argv[2])
            except ValueError:
                print(f"Invalid year: {sys.argv[2]}")
                return
        else:
            year = datetime.now().year

        print(f"Generating Arbeitnehmerveranlagung for: {person}, year: {year}")
        print(f"Fetching data from database...")
        db_rows = get_database_data(person=person, year=year)

        if not db_rows:
            print(f"No personal tax data found for {person} in {year}")
            return

        print(f"Found {len(db_rows)} total records")

        print("Aggregating tax data by month and category...")
        monthly_data = aggregate_data_by_month(db_rows)

        tax_entries = sum(len(cat['entries']) for month_data in monthly_data.values()
                          for cat in month_data.values())
        print(f"Found {tax_entries} tax-relevant entries")

        print("\nMonthly breakdown:")
        for month in sorted(monthly_data.keys()):
            month_name = MONTH_NAMES.get(month, f"Month {month}")
            total_month = sum(cat['amount'] for cat in monthly_data[month].values())
            print(f"  {month_name}: Total: {abs(total_month):.2f} €")

        output_file = f'{person.lower()}_arbeitnehmerveranlagung_{year}.xlsx'
        print(f"\nGenerating Excel file: {output_file}")
        generate_excel(monthly_data, person=person, year=year, output_file=output_file)

        print(f"✓ Excel file created successfully: {output_file}")
        print(f"\nTo download from container:")
        print(f"docker cp $(docker-compose ps -q web):/app/{output_file} ./exports/{output_file}")

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()