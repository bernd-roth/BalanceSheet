import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import sys
import os

# Database connection parameters
DB_CONFIG = {
    'host': 'HOSTNAME',
    'database': 'DATABASE_NAME',
    'user': 'USERNAME',
    'password': 'PASSWORD'
}

# Month names in German
MONTH_NAMES = {
    1: 'Jänner',
    2: 'Februar',
    3: 'März',
    4: 'April',
    5: 'Mai',
    6: 'Juni',
    7: 'Juli',
    8: 'August',
    9: 'September',
    10: 'Oktober',
    11: 'November',
    12: 'Dezember'
}

# Category mapping (using the mapped category names that match category_columns keys)
TEMPLATE_CATEGORIES = [
    'wasser/heizung',
    'haushaltsversicherung',
    'hausverwaltung',
    'mieteinkommen',
    'strom',
    'internet',
    'obs haushaltsabgabe',
    'klimaanlage',
    'rechtsschutzversicherung',
    'steuerberater',
    'bank',
    'rechts_und_beratungskosten'
]


def get_category_column(position, comment, export_to='auto'):
    """
    Determine which sheet column based on position field and export_to setting
    Returns the column name or None if not rental-related

    Args:
        position: The position/category name
        comment: Additional comment used to refine insurance categorization
        export_to: Export routing control ('auto', 'hollgasse', 'arbeitnehmerveranlagung', 'both', 'none')
    """
    # Check export_to field first
    if export_to == 'none':
        return None
    if export_to == 'arbeitnehmerveranlagung':
        return None  # This entry should only go to arbeitnehmerveranlagung, not hollgasse
    # If export_to is 'auto', 'hollgasse', or 'both', continue with position-based mapping

    position_lower = position.lower().strip() if position else ''
    comment_lower = comment.lower().strip() if comment else ''

    # Direct position mapping for rental-related expenses
    # Maps both enum names (with underscores) and display names (with slashes/spaces)
    rental_positions = {
        'mieteinkommen': 'mieteinkommen',
        'wasser/heizung': 'wasser/heizung',
        'haushaltsversicherung': 'haushaltsversicherung',
        'versicherung': 'haushaltsversicherung',  # Generic insurance maps to household insurance
        'hausverwaltung': 'hausverwaltung',
        'strom': 'strom',
        'internet': 'internet',
        'klimaanlage': 'klimaanlage',
        'obs haushaltsabgabe': 'obs haushaltsabgabe',
        'rechtsschutzversicherung': 'rechtsschutzversicherung',
        'steuerberater': 'steuerberater',
        'bank': 'bank',
        'rechts -und beratungskosten': 'rechts_und_beratungskosten',
        'rechts_und_beratungskosten': 'rechts_und_beratungskosten'
    }

    # Special handling for generic "versicherung" - check comment for specific insurance type
    if position_lower == 'versicherung':
        # Check comment for Rechtsschutzversicherung keywords
        if 'rechtsschutz' in comment_lower:
            return 'rechtsschutzversicherung'
        # Default to household insurance for other cases
        return 'haushaltsversicherung'

    # For export_to='hollgasse' or 'both', force inclusion even if not in rental_positions
    if export_to in ['hollgasse', 'both']:
        # Try to map the position, or use a default category if it doesn't match
        return rental_positions.get(position_lower, 'bank')  # Default to 'bank' for forced entries

    # For export_to='auto', check if position matches any rental category
    if position_lower in rental_positions:
        return rental_positions[position_lower]

    # Not a rental-related entry
    return None


def get_database_data(location='Hollgasse 1/54', year=2026, taxable=None):
    """Fetch rental data from Postgres database for specific location and year

    Args:
        location: The location name (e.g., 'Hollgasse 1/54')
        year: The year to filter by
        taxable: If True, filter for taxable entries. If False, filter for non-taxable entries.
                 If None, return all entries for the location.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()

    if taxable is not None:
        query = """
            SELECT id, orderdate, position, income, expense, comment, taxable, export_to
            FROM incomeexpense
            WHERE location = %s
            AND EXTRACT(YEAR FROM orderdate) = %s
            AND taxable = %s
            ORDER BY orderdate, id
        """
        cursor.execute(query, (location, year, taxable))
    else:
        query = """
            SELECT id, orderdate, position, income, expense, comment, taxable, export_to
            FROM incomeexpense
            WHERE location = %s
            AND EXTRACT(YEAR FROM orderdate) = %s
            ORDER BY orderdate, id
        """
        cursor.execute(query, (location, year))

    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    return rows


def aggregate_data_by_month(db_rows):
    """Aggregate database entries by month and category"""
    # Structure: monthly_data[month][category] = {amount: float, entries: []}
    monthly_data = defaultdict(lambda: defaultdict(lambda: {'amount': 0.0, 'entries': []}))

    for row in db_rows:
        # Handle different row formats (old and new with additional fields)
        if len(row) == 8:
            db_id, orderdate, position, income, expense, comment, taxable, export_to = row
        elif len(row) == 7:
            db_id, orderdate, position, income, expense, comment, taxable = row
            export_to = 'auto'  # Default for data without export_to field
        else:
            db_id, orderdate, position, income, expense, comment = row
            taxable = False  # Default for old data
            export_to = 'auto'

        # Determine category - skip if not rental-related
        category = get_category_column(position, comment, export_to)
        if category is None:
            continue

        month = orderdate.month
        amount = float(income) if float(income) > 0 else -float(expense)

        monthly_data[month][category]['amount'] += amount
        monthly_data[month][category]['entries'].append({
            'date': orderdate,
            'description': position,
            'amount': amount,
            'comment': comment if comment else ''
        })

    return monthly_data


def generate_excel(monthly_data, location='Hollgasse_1_54', year=2026, output_file=None):
    """Generate Excel file with proper formatting matching the template"""

    if output_file is None:
        output_file = f'{location.lower()}_{year}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = f"steuererklaerung_{location.lower()[:20]}"

    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_alignment = Alignment(horizontal="left", vertical="center")

    # Thin border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Title (merged across columns A-R) with yellow background
    ws.merge_cells('A1:S1')
    title_cell = ws['A1']
    title_cell.value = f"Einnahmen und \nAusgaben\nfür das Jahr {year}"
    title_cell.fill = yellow_fill
    title_cell.alignment = centered_alignment
    title_cell.font = Font(bold=True, size=12)
    ws.row_dimensions[1].height = 60

    # Row 2: Column headers
    headers = [
        'rechnungsnummer',
        'datum',
        'artikelbeschreibung',
        'ein / aus',
        'mieteinkommen',
        'haus-\nverwaltung',
        'haushalts-\nversicherung',
        'strom',
        'wasser/\nheizung',
        'av',
        'kleinmaterial',
        'internet',
        'klimaanlage',
        'rechtsschutz-\nversicherung',
        'steuerberater',
        'obs haushalts-\nabgabe',
        'bank',
        'rechts -und\nberatungskosten',
        'comment'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.alignment = header_alignment
        cell.font = Font(bold=True)
        cell.border = thin_border

    ws.row_dimensions[2].height = 30

    # Freeze first 2 rows
    ws.freeze_panes = 'A3'

    # Set column widths
    column_widths = {
        'A': 18,  # rechnungsnummer
        'B': 12,  # datum
        'C': 25,  # artikelbeschreibung
        'D': 12,  # ein/aus
        'E': 10,  # mieteinkommen
        'F': 12,  # hausverwaltung
        'G': 13,  # haushaltsversicherung
        'H': 10,  # strom
        'I': 10,  # wasser/heizung
        'J': 8,   # av
        'K': 13,  # kleinmaterial
        'L': 10,  # internet
        'M': 12,  # klimaanlage
        'N': 13,  # steuerberater
        'O': 13,  # obs haushaltsabgabe
        'P': 10,  # (empty/reserved)
        'Q': 10,  # bank
        'R': 14,  # rechts -und beratungskosten
        'S': 30,  # comment
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Initialize totals and invoice counter
    totals = {
        'ein_aus': 0.0,
        'mieteinkommen': 0.0,
        'hausverwaltung': 0.0,
        'haushaltsversicherung': 0.0,
        'strom': 0.0,
        'wasser/heizung': 0.0,
        'av': 0.0,
        'kleinmaterial': 0.0,
        'internet': 0.0,
        'klimaanlage': 0.0,
        'rechtsschutzversicherung': 0.0,
        'steuerberater': 0.0,
        'obs haushaltsabgabe': 0.0,
        'bank': 0.0,
        'rechts_und_beratungskosten': 0.0
    }

    invoice_number = 1
    current_row = 3

    # Column mapping
    category_columns = {
        'mieteinkommen': 5,          # E - mieteinkommen
        'hausverwaltung': 6,  # F
        'haushaltsversicherung': 7,  # G
        'strom': 8,          # H
        'wasser/heizung': 9,         # I - wasser/heizung
        'av': 10,            # J
        'kleinmaterial': 11, # K
        'internet': 12,      # L
        'klimaanlage': 13,   # M
        'rechtsschutzversicherung': 14, # N
        'steuerberater': 15, # O
        'obs haushaltsabgabe': 16, # P
        'bank': 17,          # Q - bank (for Gemeinsam tax category)
        'rechts_und_beratungskosten': 18,  # R
    }

    # Write data for each month
    for month in range(1, 13):
        month_name = MONTH_NAMES[month]

        # Month header row
        ws.cell(row=current_row, column=1).value = month_name
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        # Get categories for this month
        categories_in_month = monthly_data.get(month, {})

        # Write entries for each category (only if category has entries)
        for category in TEMPLATE_CATEGORIES:
            if category in categories_in_month:
                data = categories_in_month[category]

                # For each entry in this month/category combination
                for entry in data['entries']:
                    entry_amount = entry['amount']

                    # Column A: Invoice number (001, 002, 003...)
                    invoice_cell = ws.cell(row=current_row, column=1)
                    invoice_cell.value = invoice_number
                    invoice_cell.number_format = '000'
                    invoice_number += 1

                    # Column B: Date
                    date_cell = ws.cell(row=current_row, column=2)
                    date_cell.value = entry['date']
                    date_cell.number_format = 'dd.mm.yyyy'

                    # Column C: Description
                    ws.cell(row=current_row, column=3).value = entry['description']

                    # Column D: ein/aus (total amount)
                    ein_aus_cell = ws.cell(row=current_row, column=4)
                    ein_aus_cell.value = entry_amount
                    ein_aus_cell.number_format = '#,##0.00 [$€-1]'
                    totals['ein_aus'] += entry_amount

                    # Specific category column
                    if category in category_columns:
                        col_idx = category_columns[category]
                        amount_cell = ws.cell(row=current_row, column=col_idx)
                        amount_cell.value = entry_amount
                        amount_cell.number_format = '#,##0.00 [$€-1]'
                        totals[category] += entry_amount

                    # Column S: Comment
                    ws.cell(row=current_row, column=19).value = entry.get('comment', '')

                    current_row += 1
            # Skip empty rows - no else clause

    # Add totals row
    ws.cell(row=current_row, column=1).value = "summe"
    ws.cell(row=current_row, column=1).font = Font(bold=True)

    # Total ein/aus
    total_cell = ws.cell(row=current_row, column=4)
    total_cell.value = totals['ein_aus']
    total_cell.number_format = '#,##0.00 [$€-1]'
    total_cell.font = Font(bold=True)

    # Totals for each category
    for category, col_idx in category_columns.items():
        if totals[category] != 0:
            total_cat_cell = ws.cell(row=current_row, column=col_idx)
            total_cat_cell.value = totals[category]
            total_cat_cell.number_format = '#,##0.00 [$€-1]'
            total_cat_cell.font = Font(bold=True)

    # Save the workbook
    wb.save(output_file)

    # Fix file permissions (make it readable/writable for user and group)
    try:
        os.chmod(output_file, 0o664)
    except Exception as e:
        print(f"Warning: Could not set file permissions: {e}")

    return output_file

def main():
    """Main function to export rental data to Excel"""
    try:
        # Parse command line arguments: location [year]
        # Usage: python3 script.py Hollgasse_1_54 [2026]
        if len(sys.argv) < 2:
            print("Usage: python3 rental_export_excel.py <location> [year]")
            print("Example: python3 rental_export_excel.py Hollgasse_1_54")
            print("Example: python3 rental_export_excel.py Hollgasse_1_1 2025")
            return

        location = sys.argv[1]

        # Smart taxable detection from location name
        # Hollgasse_1_54           -> taxable=True  (default: Hollgasse 1 tax report)
        # Hollgasse_1_54_nontaxable -> taxable=False (Hollgasse 54 report)
        # Hollgasse_1_54_all       -> taxable=None  (all entries)
        taxable_filter = True  # Default to taxable=True for backward compatibility
        location_for_file = location

        if location.endswith('_nontaxable'):
            taxable_filter = False
            location = location[:-11]  # Remove '_nontaxable' suffix
            print(f"INFO: Filtering for NON-TAXABLE entries only (taxable=False)")
        elif location.endswith('_all'):
            taxable_filter = None
            location = location[:-4]  # Remove '_all' suffix
            print(f"INFO: Including ALL entries (no taxable filter)")
        else:
            print(f"INFO: Filtering for TAXABLE entries only (taxable=True) - default behavior")

        # Convert location format from command line (Hollgasse_1_1) to database format (Hollgasse 1/1)
        # Replace first underscore with space and second underscore with slash
        location_parts = location.split('_')
        if len(location_parts) == 3:  # Format: Hollgasse_1_1
            db_location = f"{location_parts[0]} {location_parts[1]}/{location_parts[2]}"
        else:
            # Use as-is if format doesn't match expected pattern
            db_location = location

        # Get year from command line argument or use current year
        if len(sys.argv) > 2:
            try:
                year = int(sys.argv[2])
            except ValueError:
                print(f"Invalid year: {sys.argv[2]}")
                print("Usage: python3 rental_export_excel.py <location> [year]")
                print("Examples:")
                print("  python3 hollgasse.py Hollgasse_1_54              # Taxable entries (default)")
                print("  python3 hollgasse.py Hollgasse_1_54_nontaxable   # Non-taxable entries")
                print("  python3 hollgasse.py Hollgasse_1_54_all          # All entries")
                return
        else:
            # Use current year based on system timestamp
            year = datetime.now().year

        # Allow explicit override via 3rd parameter (optional, for manual testing)
        if len(sys.argv) > 3:
            taxable_arg = sys.argv[3].lower()
            if taxable_arg == 'true':
                taxable_filter = True
                print(f"INFO: Explicit override - taxable=True")
            elif taxable_arg == 'false':
                taxable_filter = False
                print(f"INFO: Explicit override - taxable=False")
            elif taxable_arg == 'all' or taxable_arg == 'none':
                taxable_filter = None
                print(f"INFO: Explicit override - taxable=None (all entries)")

        print(f"Generating report for location: {location}, year: {year}, taxable: {taxable_filter}")
        print(f"Database location format: {db_location}")
        print(f"Fetching data from database...")
        db_rows = get_database_data(location=db_location, year=year, taxable=taxable_filter)

        if not db_rows:
            print(f"No data found for {year}")
            return

        print(f"Found {len(db_rows)} total records")

        # Debug: Show taxable status and export_to of fetched entries
        print("\nDEBUG - Fetched entries with taxable and export_to status:")
        for row in db_rows[:10]:  # Show first 10 entries
            if len(row) == 8:
                db_id, orderdate, position, income, expense, comment, taxable, export_to = row
                print(f"  ID {db_id}: {position} | {expense}€ | taxable={taxable} | export_to={export_to}")
            elif len(row) == 7:
                db_id, orderdate, position, income, expense, comment, taxable = row
                print(f"  ID {db_id}: {position} | {expense}€ | taxable={taxable} | export_to=MISSING")
            else:
                db_id, orderdate, position, income, expense, comment = row
                print(f"  ID {db_id}: {position} | {expense}€ | taxable=MISSING | export_to=MISSING")

        print("Aggregating rental-related data by month and category...")
        monthly_data = aggregate_data_by_month(db_rows)

        rental_entries = sum(len(cat['entries']) for month_data in monthly_data.values()
                             for cat in month_data.values())
        print(f"Found {rental_entries} rental-related entries")

        print("\nMonthly breakdown:")
        for month in sorted(monthly_data.keys()):
            month_name = MONTH_NAMES.get(month, f"Month {month}")
            categories = list(monthly_data[month].keys())
            total_month = sum(cat['amount'] for cat in monthly_data[month].values())
            print(f"  {month_name}: {len(categories)} categories, Total: {total_month:.2f} €")

        # Use the original location (with suffix) for filename
        output_file = f'{location_for_file.lower()}_{year}.xlsx'
        print(f"\nGenerating Excel file: {output_file}")
        generate_excel(monthly_data, location=location_for_file, year=year, output_file=output_file)

        print(f"✓ Excel file created successfully: {output_file}")
        print(f"\nThe file includes:")
        print(f"  ✓ Location: {location}")
        print(f"  ✓ Year: {year}")
        print(f"  ✓ Yellow header with centered text")
        print(f"  ✓ Frozen first 2 rows")
        print(f"  ✓ Sequential invoice numbering (001, 002, 003...)")
        print(f"  ✓ Proper date and currency formatting")
        print(f"  ✓ Monthly breakdown with totals")
        print(f"\nTo download from container:")
        print(f"docker cp $(docker-compose ps -q web):/app/{output_file} ./exports/{output_file}")

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()